import csv
import logging
import re
from datetime import date, datetime
from typing import Optional

import openpyxl
from sqlmodel import Session, select

logger = logging.getLogger(__name__)

from models import (
    DbSourcePatent, DbSourceInventor,
    UnifiedPatent, UnifiedInventor,
    PatentCrossRef, ReconciliationChoice,
    InventorAttendance, AwardCost,
)


def normalize_patent_number(raw: str) -> str:
    """Extract numeric patent identifier, keeping D prefix for design patents."""
    cleaned = raw.strip().upper().replace(" ", "")
    # Remove US prefix (with or without hyphen)
    cleaned = re.sub(r"^US-?", "", cleaned)
    # Remove suffix like -B2, -S1, -A1
    cleaned = re.sub(r"-[A-Z]\d*$", "", cleaned)
    return cleaned


def normalize_name(name: str) -> str:
    """Lowercase, split tokens, sort alphabetically."""
    return " ".join(sorted(name.strip().lower().split()))


def parse_date(val) -> Optional[date]:
    """Parse various date formats to date object."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d", "%m/%d/%y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def clear_db_source(session: Session, project_id: int):
    """Delete existing db_source data for a project."""
    patents = session.exec(
        select(DbSourcePatent).where(DbSourcePatent.project_id == project_id)
    ).all()
    for p in patents:
        inventors = session.exec(
            select(DbSourceInventor).where(DbSourceInventor.db_source_patent_id == p.id)
        ).all()
        for inv in inventors:
            session.delete(inv)
        session.delete(p)
    # Also clear crossrefs since data changed
    _clear_crossrefs(session, project_id)
    session.commit()


def clear_unified(session: Session, project_id: int):
    """Delete existing unified data for a project."""
    patents = session.exec(
        select(UnifiedPatent).where(UnifiedPatent.project_id == project_id)
    ).all()
    for p in patents:
        inventors = session.exec(
            select(UnifiedInventor).where(UnifiedInventor.unified_patent_id == p.id)
        ).all()
        for inv in inventors:
            session.delete(inv)
        session.delete(p)
    _clear_crossrefs(session, project_id)
    session.commit()


def _clear_crossrefs(session: Session, project_id: int):
    crossrefs = session.exec(
        select(PatentCrossRef).where(PatentCrossRef.project_id == project_id)
    ).all()
    for cr in crossrefs:
        choices = session.exec(
            select(ReconciliationChoice).where(ReconciliationChoice.crossref_id == cr.id)
        ).all()
        for ch in choices:
            session.delete(ch)
        session.delete(cr)


def import_db_source(session: Session, project_id: int, file_path: str) -> dict:
    """Import db_source CSV. Groups rows by Patent No. to create one patent with multiple inventors."""
    clear_db_source(session, project_id)

    # Try UTF-8 first, fall back to latin-1 (accepts all byte values)
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            with open(file_path, "r", encoding=encoding) as f:
                reader = csv.DictReader(f)
                rows = list(reader)
                headers = reader.fieldnames
            logger.info("CSV opened with encoding=%s, headers: %s", encoding, headers)
            break
        except UnicodeDecodeError:
            logger.warning("Failed to read CSV with encoding=%s, trying next", encoding)
            continue

    logger.info("Read %d rows from %s", len(rows), file_path)

    # Group by Patent No.
    groups: dict[str, list[dict]] = {}
    for i, row in enumerate(rows):
        patent_no_val = row.get("Patent No.")
        if patent_no_val is None:
            logger.warning("Row %d missing 'Patent No.' column. Keys: %s", i + 1, list(row.keys()))
            continue
        key = patent_no_val.strip()
        if not key:
            logger.warning("Row %d has empty 'Patent No.' value, skipping", i + 1)
            continue
        groups.setdefault(key, []).append(row)

    logger.info("Grouped into %d unique patents", len(groups))

    patent_count = 0
    inventor_count = 0

    for patent_no, group_rows in groups.items():
        first = group_rows[0]
        try:
            patent = DbSourcePatent(
                project_id=project_id,
                asset_name=first["Patent: Asset Name"].strip(),
                patent_no=patent_no,
                patent_no_numeric=normalize_patent_number(patent_no),
                title=first["Title"].strip(),
                issue_date=parse_date(first["Issue Date of Patent"]),
            )
            session.add(patent)
            session.flush()  # get patent.id
            patent_count += 1
        except Exception:
            logger.exception("Failed to create patent from row with Patent No. '%s'. Row data: %s", patent_no, first)
            raise

        for row in group_rows:
            try:
                name = row["Inventor: Legal Name"].strip()
                inventor = DbSourceInventor(
                    db_source_patent_id=patent.id,
                    legal_name=name,
                    name_normalized=normalize_name(name),
                    office_location_country=row.get("Inventor: Office Location Country", "").strip() or None,
                    work_country_iso=row.get("Inventor: Work Country ISO code", "").strip() or None,
                    address=row.get("Inventor: Address", "").strip() or None,
                    award_type=row.get("Inventor: Award Type", "").strip() or None,
                    work_city=row.get("Inventor: Work City", "").strip() or None,
                    work_state=row.get("Inventor: Work State", "").strip() or None,
                    work_email=row.get("Inventor: Work Email", "").strip() or None,
                    preferred_name=row.get("Inventor: Preferred Name", "").strip() or None,
                    employment_status=row.get("Inventor: Employment Status", "").strip() or None,
                    employee_id=row.get("Inventor: Employee ID", "").strip() or None,
                )
                session.add(inventor)
                inventor_count += 1
            except Exception:
                logger.exception("Failed to create inventor from row: %s", row)
                raise

    # Auto-create AwardCost rows for new award types
    seen_award_types: set[str] = set()
    for group_rows in groups.values():
        for row in group_rows:
            at = row.get("Inventor: Award Type", "").strip()
            if at and at.lower() != "opt-out":
                seen_award_types.add(at)

    existing_costs = session.exec(
        select(AwardCost).where(AwardCost.project_id == project_id)
    ).all()
    existing_types = {c.award_type for c in existing_costs}
    costs_created = 0
    for at in sorted(seen_award_types):
        if at not in existing_types:
            session.add(AwardCost(project_id=project_id, award_type=at, cost=0.0))
            costs_created += 1

    session.commit()
    logger.info("db_source import complete: %d patents, %d inventors, %d new award cost rows", patent_count, inventor_count, costs_created)
    return {"patents_imported": patent_count, "inventors_imported": inventor_count, "award_costs_created": costs_created}


def import_unified(session: Session, project_id: int, file_path: str) -> dict:
    """Import unified XLSX. Splits pipe-delimited inventors into separate rows."""
    clear_unified(session, project_id)

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Read header row
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    logger.info("XLSX headers: %s", headers)

    patent_count = 0
    inventor_count = 0
    skipped = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = dict(zip(headers, row))
        pub_number = str(data.get("publication_number", "")).strip()
        if not pub_number:
            logger.warning("Row %d: empty publication_number, skipping. Row data: %s", row_idx, data)
            skipped += 1
            continue

        try:
            patent = UnifiedPatent(
                project_id=project_id,
                publication_number=pub_number,
                patent_no_numeric=normalize_patent_number(pub_number),
                publication_date=parse_date(data.get("publication_date")),
                title=str(data.get("title", "")).strip(),
                grant_number=str(data.get("grant_number", "")).strip(),
                assignee_current=str(data.get("assignee_current", "")).strip() or None,
                assignee_original=str(data.get("assignee_original", "")).strip() or None,
                assignee_parent=str(data.get("assignee_parent", "")).strip() or None,
                application_date=parse_date(data.get("application_date")),
                priority_date=parse_date(data.get("priority_date")),
                publication_status=str(data.get("publication_status", "")).strip() or None,
                publication_type=str(data.get("publication_type", "")).strip() or None,
                country=str(data.get("country", "")).strip() or None,
            )
            session.add(patent)
            session.flush()
            patent_count += 1
        except Exception:
            logger.exception("Row %d: failed to create unified patent. pub_number='%s', data=%s", row_idx, pub_number, data)
            raise

        # Split inventors by pipe
        inventors_raw = str(data.get("inventors", "")).strip()
        if inventors_raw:
            for name in inventors_raw.split("|"):
                name = name.strip()
                if name:
                    inv = UnifiedInventor(
                        unified_patent_id=patent.id,
                        raw_name=name,
                        name_normalized=normalize_name(name),
                    )
                    session.add(inv)
                    inventor_count += 1

    logger.info("unified import complete: %d patents, %d inventors, %d rows skipped", patent_count, inventor_count, skipped)
    session.commit()
    return {"patents_imported": patent_count, "inventors_imported": inventor_count}


def import_attendance(session: Session, project_id: int, file_path: str) -> dict:
    """Import attendance CSV. Columns: employee_id, email. Sets status to Unknown."""
    # Clear existing attendance for project
    existing = session.exec(
        select(InventorAttendance).where(InventorAttendance.project_id == project_id)
    ).all()
    for row in existing:
        session.delete(row)
    session.flush()

    # Read CSV
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            with open(file_path, "r", encoding=encoding) as f:
                reader = csv.DictReader(f)
                rows = list(reader)
            break
        except UnicodeDecodeError:
            continue

    count = 0
    for row in rows:
        emp_id = row.get("employee_id", "").strip()
        email = row.get("email", "").strip()
        if not emp_id and not email:
            continue
        att = InventorAttendance(
            project_id=project_id,
            employee_id=emp_id,
            email=email,
            attendance_status="Unknown",
        )
        session.add(att)
        count += 1

    session.commit()
    logger.info("attendance import complete: %d records", count)
    return {"imported": count}

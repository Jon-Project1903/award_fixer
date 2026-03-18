import io
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from sqlmodel import Session, select

from models import (
    PhysicalAward, AwardCost, TaxRate, ProgramMgmtFee,
    ShippingAddress, InventorShipping,
    PatentCrossRef, DbSourcePatent, ReconciliationChoice,
)
from services.award_service import compute_cost_summary

HEADER_FONT = Font(bold=True, color="333333")
HEADER_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
TOTAL_FONT = Font(bold=True, color="1E3A5F")
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBTOTAL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
CURRENCY_FMT = '"$"#,##0.00'
PERCENT_FMT = '0.00"%"'


def _style_header(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def _format_address(addr):
    if not addr:
        return "Unknown"
    parts = [addr.get("address_1") or "", addr.get("city") or "", addr.get("state") or "",
             addr.get("zip_code") or "", addr.get("country") or ""]
    return ", ".join(p for p in parts if p)


def generate_report(session: Session, project_id: int) -> io.BytesIO:
    wb = Workbook()

    # --- Sheet 1: Budget Estimates ---
    ws = wb.active
    ws.title = "Budget Estimates"
    _build_budget_sheet(ws, session, project_id)

    # --- Sheet 2: Shipping Addresses ---
    ws2 = wb.create_sheet("Shipping Addresses")
    _build_shipping_sheet(ws2, session, project_id)

    # --- Sheet 3: Tax Rates ---
    ws3 = wb.create_sheet("Tax Rates")
    _build_tax_rates_sheet(ws3, session, project_id)

    # --- Sheet 4: All Awards ---
    awards_data = _load_awards_data(session, project_id)
    ws4 = wb.create_sheet("All Awards")
    _build_awards_sheet(ws4, awards_data)

    # --- Sheets 5+: Per award type ---
    by_type = defaultdict(list)
    for row in awards_data:
        by_type[row["award_type"]].append(row)

    for award_type in sorted(by_type.keys()):
        sheet_name = award_type[:31]  # Excel 31-char limit
        ws_type = wb.create_sheet(sheet_name)
        _build_awards_sheet(ws_type, by_type[award_type])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_budget_sheet(ws, session, project_id):
    summary = compute_cost_summary(session, project_id)

    headers = ["", "Qty", "Price Each", "Total"]
    ws.append(headers)
    _style_header(ws, len(headers))

    # Line items
    for item in summary["line_items"]:
        ws.append([item["award_type"], item["quantity"], item["unit_cost"], item["total"]])

    # Subtotal
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value="Subtotal").font = Font(bold=True)
    ws.cell(row=row, column=4, value=summary["subtotal"]).font = Font(bold=True)
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = SUBTOTAL_FILL

    # Tax breakdown
    for tb in summary["tax_breakdown"]:
        ws.append([f"  Tax: {tb['jurisdiction']}", f"{tb['rate']}%", f"on {tb['taxable_amount']:.2f}", tb["tax"]])

    # Total tax
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value="TAXES estimated").font = Font(bold=True)
    ws.cell(row=row, column=4, value=summary["total_tax"]).font = Font(bold=True)

    # Subtotal incl tax
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value="Subtotal incl. tax").font = Font(bold=True)
    ws.cell(row=row, column=4, value=summary["subtotal_with_tax"]).font = Font(bold=True)
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = SUBTOTAL_FILL

    # Other fees
    for fee in summary["pm_fees"]:
        ws.append([fee["description"], fee["quantity"], fee["unit_cost"], fee["total"]])

    # Grand total
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value="GRAND TOTAL").font = TOTAL_FONT
    ws.cell(row=row, column=4, value=summary["grand_total"]).font = TOTAL_FONT
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = TOTAL_FILL

    # Format currency columns
    for row_cells in ws.iter_rows(min_row=2, min_col=3, max_col=4):
        for cell in row_cells:
            if isinstance(cell.value, (int, float)):
                cell.number_format = CURRENCY_FMT

    _auto_width(ws)


def _build_shipping_sheet(ws, session, project_id):
    addresses = session.exec(
        select(ShippingAddress).where(ShippingAddress.project_id == project_id)
    ).all()

    headers = ["City", "State", "Country", "Company", "Ship To", "Email",
               "Address 1", "Address 2", "Address 3", "Phone", "Zip Code", "Taxable"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for addr in addresses:
        ws.append([
            addr.city, addr.state, addr.country, addr.company, addr.ship_to, addr.email,
            addr.address_1, addr.address_2, addr.address_3, addr.phone, addr.zip_code,
            "Yes" if addr.taxable else "No",
        ])

    _auto_width(ws)


def _build_tax_rates_sheet(ws, session, project_id):
    rates = session.exec(
        select(TaxRate).where(TaxRate.project_id == project_id)
    ).all()

    headers = ["City", "State", "Tax %"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for rate in rates:
        row_num = ws.max_row + 1
        ws.cell(row=row_num, column=1, value=rate.jurisdiction)
        ws.cell(row=row_num, column=2, value=rate.lookup_key)
        cell = ws.cell(row=row_num, column=3, value=rate.tax_percent)
        cell.number_format = '0.00'

    _auto_width(ws)


def _load_awards_data(session, project_id):
    awards = session.exec(
        select(PhysicalAward).where(PhysicalAward.project_id == project_id)
    ).all()

    # Cost lookup
    costs = session.exec(
        select(AwardCost).where(AwardCost.project_id == project_id)
    ).all()
    cost_by_type = {c.award_type: c.cost for c in costs}

    # Tax lookup by city
    tax_rates = session.exec(
        select(TaxRate).where(TaxRate.project_id == project_id)
    ).all()
    tax_by_city = {r.jurisdiction.strip().lower(): r for r in tax_rates}

    # Shipping lookup by employee_id
    shipping_rows = session.exec(
        select(InventorShipping).where(InventorShipping.project_id == project_id)
    ).all()
    shipping_by_emp = {}
    for s in shipping_rows:
        shipping_by_emp[s.employee_id.strip().lower()] = s

    # Shipping addresses for formatting
    addresses = session.exec(
        select(ShippingAddress).where(ShippingAddress.project_id == project_id)
    ).all()
    addr_by_id = {a.id: a for a in addresses}

    # Build patent_number -> reconciled title lookup
    crossrefs = session.exec(
        select(PatentCrossRef).where(
            PatentCrossRef.project_id == project_id,
            PatentCrossRef.db_source_patent_id != None,
        )
    ).all()
    title_by_patent_no = {}
    for cr in crossrefs:
        db_pat = session.get(DbSourcePatent, cr.db_source_patent_id)
        if not db_pat:
            continue
        # Check for a reconciled title choice
        title_choice = session.exec(
            select(ReconciliationChoice).where(
                ReconciliationChoice.crossref_id == cr.id,
                ReconciliationChoice.field_name == "title",
            )
        ).first()
        if title_choice:
            title_by_patent_no[db_pat.patent_no] = title_choice.chosen_value
        else:
            title_by_patent_no[db_pat.patent_no] = db_pat.title

    rows = []
    for a in awards:
        unit_cost = cost_by_type.get(a.award_type, 0.0)
        city = (a.work_city or "").strip()
        tax_rate = 0.0
        if city and city.lower() in tax_by_city:
            tax_rate = tax_by_city[city.lower()].tax_percent
        tax_amount = round(unit_cost * (tax_rate / 100.0), 2)

        # Delivery address
        emp_key = (a.employee_id or "").strip().lower()
        shipping = shipping_by_emp.get(emp_key)
        if shipping and shipping.shipping_type == "To the Event":
            delivery = "To the Event"
        elif shipping and shipping.shipping_address_id:
            addr = addr_by_id.get(shipping.shipping_address_id)
            if addr:
                parts = [addr.address_1, addr.city, addr.state, addr.zip_code, addr.country]
                delivery = ", ".join(p for p in parts if p)
            else:
                delivery = "Unknown"
        else:
            delivery = "Unknown"

        rows.append({
            "inventor_name": a.inventor_name,
            "employee_id": a.employee_id,
            "patent_number": a.patent_number,
            "patent_title": title_by_patent_no.get(a.patent_number, ""),
            "award_type": a.award_type,
            "unit_cost": unit_cost,
            "work_city": city,
            "tax_rate": tax_rate,
            "tax_amount": tax_amount,
            "delivery": delivery,
        })

    return rows


def _build_awards_sheet(ws, rows):
    headers = ["Inventor Name", "Employee ID", "Patent Number", "Patent Title", "Award Type",
               "Unit Cost", "Work City", "Tax Rate (%)", "Tax Amount ($)", "Delivery Address"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for r in rows:
        row_num = ws.max_row + 1
        ws.cell(row=row_num, column=1, value=r["inventor_name"])
        ws.cell(row=row_num, column=2, value=r["employee_id"])
        ws.cell(row=row_num, column=3, value=r["patent_number"])
        ws.cell(row=row_num, column=4, value=r["patent_title"])
        ws.cell(row=row_num, column=5, value=r["award_type"])
        ws.cell(row=row_num, column=6, value=r["unit_cost"]).number_format = CURRENCY_FMT
        ws.cell(row=row_num, column=7, value=r["work_city"])
        ws.cell(row=row_num, column=8, value=r["tax_rate"]).number_format = '0.00'
        ws.cell(row=row_num, column=9, value=r["tax_amount"]).number_format = CURRENCY_FMT
        ws.cell(row=row_num, column=10, value=r["delivery"])

    _auto_width(ws)
